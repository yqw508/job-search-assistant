from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter

from boss_job_assistant.models import ScoredJob


HEADERS = [
    "匹配状态",
    "匹配分",
    "岗位名称",
    "薪资",
    "地点",
    "经验",
    "学历",
    "公司",
    "行业",
    "融资阶段",
    "公司规模",
    "岗位链接",
    "命中原因",
    "排除原因",
    "岗位描述摘要",
]


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _clean_row(values: list[Any]) -> list[Any]:
    return [_clean_cell(value) for value in values]


def export_jobs(jobs: list[ScoredJob], output_dir: str | Path) -> Path:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    filename = f"boss_jobs_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
    file_path = output_path / filename

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Boss岗位"
    sheet.append(_clean_row(HEADERS))

    for scored_job in sorted(jobs, key=lambda item: item.score, reverse=True):
        job = scored_job.job
        sheet.append(
            _clean_row(
                [
                    "匹配" if scored_job.matched else "排除",
                    scored_job.score,
                    job.title,
                    job.salary,
                    job.location,
                    job.experience,
                    job.education,
                    job.company,
                    job.industry,
                    job.financing,
                    job.company_size,
                    job.url,
                    "；".join(scored_job.matched_reasons),
                    scored_job.exclusion_reason,
                    job.description[:300],
                ]
            )
        )

    sheet.auto_filter.ref = sheet.dimensions
    widths = [10, 10, 18, 12, 12, 12, 10, 18, 14, 12, 14, 36, 28, 28, 48]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(file_path)
    return file_path
