from openpyxl import load_workbook

from boss_job_assistant.exporter import export_jobs
from boss_job_assistant.models import JobPosting, ScoredJob


def make_scored_job(
    title: str = "Java 后端",
    score: int = 90,
    matched: bool = True,
    matched_reasons: list[str] | None = None,
    exclusion_reason: str = "",
    description: str = "Java Spring Boot C端交易系统",
) -> ScoredJob:
    job = JobPosting(
        title=title,
        salary="25-35K",
        location="广州",
        experience="5-10年",
        education="本科",
        company="示例公司",
        industry="电商",
        financing="B轮",
        company_size="100-499人",
        url="https://www.zhipin.com/job_detail/example.html",
        description=description,
    )
    return ScoredJob(
        job=job,
        matched=matched,
        score=score,
        matched_reasons=matched_reasons or ["技术关键词: Java"],
        exclusion_reason=exclusion_reason,
    )


def test_export_jobs_writes_excel_with_expected_columns(tmp_path):
    scored = make_scored_job()

    file_path = export_jobs([scored], tmp_path)

    assert file_path.exists()
    workbook = load_workbook(file_path)
    sheet = workbook.active
    assert sheet["A1"].value == "匹配状态"
    assert sheet["C2"].value == "Java 后端"
    assert sheet["M2"].value == "技术关键词: Java"


def test_export_jobs_consecutive_calls_create_different_file_paths(tmp_path):
    first_path = export_jobs([make_scored_job()], tmp_path)
    second_path = export_jobs([make_scored_job()], tmp_path)

    assert first_path != second_path
    assert first_path.exists()
    assert second_path.exists()


def test_export_jobs_removes_illegal_characters_from_text_cells(tmp_path):
    scored = make_scored_job(
        title="Java\x01 后端",
        description="Java Spring\x02 Boot C端交易系统",
    )

    file_path = export_jobs([scored], tmp_path)

    workbook = load_workbook(file_path)
    sheet = workbook.active
    assert sheet["C2"].value == "Java 后端"
    assert sheet["O2"].value == "Java Spring Boot C端交易系统"


def test_export_jobs_sorts_by_score_descending(tmp_path):
    lower_score = make_scored_job(title="低分岗位", score=60)
    higher_score = make_scored_job(title="高分岗位", score=95)

    file_path = export_jobs([lower_score, higher_score], tmp_path)

    workbook = load_workbook(file_path)
    sheet = workbook.active
    assert sheet["C2"].value == "高分岗位"
    assert sheet["B2"].value == 95
    assert sheet["C3"].value == "低分岗位"
    assert sheet["B3"].value == 60


def test_export_jobs_writes_excluded_status_and_reason(tmp_path):
    scored = make_scored_job(
        matched=False,
        matched_reasons=[],
        exclusion_reason="地点不匹配",
    )

    file_path = export_jobs([scored], tmp_path)

    workbook = load_workbook(file_path)
    sheet = workbook.active
    assert sheet["A2"].value == "排除"
    assert sheet["N2"].value == "地点不匹配"
